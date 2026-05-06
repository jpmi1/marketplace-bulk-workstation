from __future__ import annotations

import json
import sqlite3
from io import BytesIO, StringIO
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .validation import sanitize_public_description, validate_listing


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PROJECT_DIR = ROOT / "projects" / "default"
DEFAULT_DB_PATH = DEFAULT_PROJECT_DIR / "marketplace.db"


DEFAULT_SETTINGS: dict[str, Any] = {
    "project_name": "Sell to 1 BTC",
    "location": "Your City, ST ZIP",
    "default_condition": "Used - Good",
    "default_payment_terms": "Cash, Venmo, or Zelle accepted.",
    "default_pickup_terms": "Local pickup available.",
    "shipping_enabled_default": True,
    "default_package_weight_oz": 16,
    "carrier_preference": "Facebook shipping / buyer-paid actual cost",
    "auto_publish": False,
    "draft_and_confirm": True,
    "batch_size": 50,
    "facebook_profile_path": str((ROOT / "outputs" / "facebook_browser_profile").resolve()),
    "image_research_enabled": False,
    "comp_research_enabled": False,
    "reference_image_policy": "review_required",
    "description_tone": "concise_buyer_facing",
    "btc_goal_amount": 1.0,
    "btc_owned": 0.0,
    "manual_btc_usd_price": 100000,
    "kraken_referral_url": "",
    "google_sheet_url": "",
    "progress_currency": "USD",
    "forbidden_public_phrases": [
        "storage inventory",
        "internal",
        "pipeline",
        "notes from sheet",
        "notes from inventory sheet",
        "inventory condition mix",
        "photos still need",
        "photo needed",
        "reference/product photos are included",
        "untested by pipeline",
    ],
}


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def connect(db_path: Path = DEFAULT_DB_PATH) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("pragma foreign_keys = on")
    return conn


def init_db(db_path: Path = DEFAULT_DB_PATH) -> None:
    with connect(db_path) as conn:
        conn.executescript(
            """
            create table if not exists settings (
              key text primary key,
              value text not null
            );
            create table if not exists listings (
              id text primary key,
              source text not null default 'manual',
              title text not null default '',
              price integer,
              condition text not null default 'Used - Good',
              category text not null default '',
              quantity_text text not null default '',
              description text not null default '',
              location text not null default '',
              pickup_enabled integer not null default 1,
              shipping_enabled integer not null default 1,
              package_weight_oz real,
              carrier_preference text not null default '',
              tags text not null default '',
              private_notes text not null default '',
              approved integer not null default 0,
              status text not null default 'needs_review',
              reference_only_approved integer not null default 0,
              comps_json text not null default '[]',
              validation_json text not null default '[]',
              posting_status text not null default '',
              posted_url text not null default '',
              created_at text not null,
              updated_at text not null
            );
            create table if not exists photos (
              id text primary key,
              listing_id text not null,
              uri text not null,
              path text not null default '',
              source_url text not null default '',
              kind text not null default 'original',
              provenance text not null default '',
              selected integer not null default 1,
              removed integer not null default 0,
              cover integer not null default 0,
              sort_order integer not null default 0,
              rights_warning text not null default '',
              created_at text not null,
              updated_at text not null,
              foreign key (listing_id) references listings(id) on delete cascade
            );
            create table if not exists run_log (
              id integer primary key autoincrement,
              listing_id text,
              level text not null,
              message text not null,
              details_json text not null default '{}',
              created_at text not null
            );
            create table if not exists btc_progress_entries (
              id integer primary key autoincrement,
              entry_type text not null,
              title text not null default '',
              amount_usd real not null default 0,
              btc_amount real not null default 0,
              btc_price_usd real,
              listing_id text,
              notes text not null default '',
              entry_date text not null,
              created_at text not null,
              updated_at text not null
            );
            """
        )
        existing = {row["key"] for row in conn.execute("select key from settings")}
        for key, value in DEFAULT_SETTINGS.items():
            if key not in existing:
                conn.execute("insert into settings(key, value) values(?, ?)", (key, json.dumps(value)))
        project_name = conn.execute("select value from settings where key = 'project_name'").fetchone()
        if project_name and json.loads(project_name["value"]) == "Marketplace Bulk Listing Workstation":
            conn.execute("update settings set value = ? where key = 'project_name'", (json.dumps(DEFAULT_SETTINGS["project_name"]),))


def get_settings(db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    init_db(db_path)
    with connect(db_path) as conn:
        settings = dict(DEFAULT_SETTINGS)
        for row in conn.execute("select key, value from settings"):
            settings[row["key"]] = json.loads(row["value"])
        return settings


def update_settings(patch: dict[str, Any], db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    init_db(db_path)
    allowed = set(DEFAULT_SETTINGS)
    with connect(db_path) as conn:
        for key, value in patch.items():
            if key not in allowed:
                continue
            conn.execute(
                "insert into settings(key, value) values(?, ?) on conflict(key) do update set value=excluded.value",
                (key, json.dumps(value)),
            )
    revalidate_all(db_path)
    return get_settings(db_path)


def row_to_listing(row: sqlite3.Row, photos: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "id": row["id"],
        "source": row["source"],
        "title": row["title"],
        "price": row["price"],
        "condition": row["condition"],
        "category": row["category"],
        "quantity_text": row["quantity_text"],
        "description": row["description"],
        "location": row["location"],
        "pickup_enabled": bool(row["pickup_enabled"]),
        "shipping_enabled": bool(row["shipping_enabled"]),
        "package_weight_oz": row["package_weight_oz"],
        "carrier_preference": row["carrier_preference"],
        "tags": json.loads(row["tags"] or "[]") if row["tags"].startswith("[") else [],
        "private_notes": row["private_notes"],
        "approved": bool(row["approved"]),
        "status": row["status"],
        "reference_only_approved": bool(row["reference_only_approved"]),
        "comps": json.loads(row["comps_json"] or "[]"),
        "validation": json.loads(row["validation_json"] or "[]"),
        "posting_status": row["posting_status"],
        "posted_url": row["posted_url"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "photos": photos,
    }


def row_to_photo(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "listing_id": row["listing_id"],
        "uri": row["uri"],
        "path": row["path"],
        "source_url": row["source_url"],
        "kind": row["kind"],
        "provenance": row["provenance"],
        "selected": bool(row["selected"]),
        "removed": bool(row["removed"]),
        "cover": bool(row["cover"]),
        "sort_order": row["sort_order"],
        "rights_warning": row["rights_warning"],
    }


def list_listings(status: str | None = None, db_path: Path = DEFAULT_DB_PATH) -> list[dict[str, Any]]:
    init_db(db_path)
    with connect(db_path) as conn:
        params: list[Any] = []
        where = ""
        if status and status != "all":
            where = "where status = ?"
            params.append(status)
        rows = conn.execute(f"select * from listings {where} order by updated_at desc, id", params).fetchall()
        return [row_to_listing(row, list_photos_for_listing(conn, row["id"])) for row in rows]


def get_listing(listing_id: str, db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any] | None:
    init_db(db_path)
    with connect(db_path) as conn:
        row = conn.execute("select * from listings where id = ?", (listing_id,)).fetchone()
        if not row:
            return None
        return row_to_listing(row, list_photos_for_listing(conn, listing_id))


def list_photos_for_listing(conn: sqlite3.Connection, listing_id: str) -> list[dict[str, Any]]:
    rows = conn.execute("select * from photos where listing_id = ? order by sort_order, id", (listing_id,)).fetchall()
    return [row_to_photo(row) for row in rows]


def upsert_listing(data: dict[str, Any], db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    init_db(db_path)
    settings = get_settings(db_path)
    listing_id = str(data["id"])
    timestamp = now()
    description = sanitize_public_description(str(data.get("description") or ""), settings.get("forbidden_public_phrases"))
    with connect(db_path) as conn:
        existing = conn.execute("select created_at from listings where id = ?", (listing_id,)).fetchone()
        conn.execute(
            """
            insert into listings(
              id, source, title, price, condition, category, quantity_text, description, location,
              pickup_enabled, shipping_enabled, package_weight_oz, carrier_preference, tags, private_notes,
              approved, status, reference_only_approved, comps_json, validation_json, posting_status, posted_url,
              created_at, updated_at
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            on conflict(id) do update set
              source=excluded.source, title=excluded.title, price=excluded.price, condition=excluded.condition,
              category=excluded.category, quantity_text=excluded.quantity_text, description=excluded.description,
              location=excluded.location, pickup_enabled=excluded.pickup_enabled, shipping_enabled=excluded.shipping_enabled,
              package_weight_oz=excluded.package_weight_oz, carrier_preference=excluded.carrier_preference,
              tags=excluded.tags, private_notes=excluded.private_notes, approved=excluded.approved, status=excluded.status,
              reference_only_approved=excluded.reference_only_approved, comps_json=excluded.comps_json,
              posting_status=excluded.posting_status, posted_url=excluded.posted_url,
              updated_at=excluded.updated_at
            """,
            (
                listing_id,
                data.get("source", "manual"),
                str(data.get("title") or "")[:150],
                data.get("price"),
                data.get("condition") or settings["default_condition"],
                data.get("category") or "",
                data.get("quantity_text") or "",
                description,
                data.get("location") or settings["location"],
                int(bool(data.get("pickup_enabled", True))),
                int(bool(data.get("shipping_enabled", settings["shipping_enabled_default"]))),
                data.get("package_weight_oz") or None,
                data.get("carrier_preference") or settings["carrier_preference"],
                json.dumps(data.get("tags") or []),
                data.get("private_notes") or "",
                int(bool(data.get("approved", False))),
                data.get("status") or "needs_review",
                int(bool(data.get("reference_only_approved", False))),
                json.dumps(data.get("comps") or []),
                "[]",
                data.get("posting_status") or "",
                data.get("posted_url") or "",
                existing["created_at"] if existing else timestamp,
                timestamp,
            ),
        )
        for photo in data.get("photos") or []:
            upsert_photo(conn, listing_id, photo)
    refresh_validation(listing_id, db_path)
    return get_listing(listing_id, db_path) or {}


def upsert_photo(conn: sqlite3.Connection, listing_id: str, photo: dict[str, Any]) -> None:
    timestamp = now()
    photo_id = str(photo.get("id") or f"{listing_id}-{photo.get('sort_order', 0)}")
    existing = conn.execute("select created_at from photos where id = ?", (photo_id,)).fetchone()
    conn.execute(
        """
        insert into photos(id, listing_id, uri, path, source_url, kind, provenance, selected, removed, cover, sort_order, rights_warning, created_at, updated_at)
        values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        on conflict(id) do update set
          uri=excluded.uri, path=excluded.path, source_url=excluded.source_url, kind=excluded.kind,
          provenance=excluded.provenance, selected=excluded.selected, removed=excluded.removed,
          cover=excluded.cover, sort_order=excluded.sort_order, rights_warning=excluded.rights_warning,
          updated_at=excluded.updated_at
        """,
        (
            photo_id,
            listing_id,
            photo.get("uri") or photo.get("path") or photo.get("source_url") or "",
            photo.get("path") or "",
            photo.get("source_url") or "",
            photo.get("kind") or "original",
            photo.get("provenance") or "",
            int(bool(photo.get("selected", True))),
            int(bool(photo.get("removed", False))),
            int(bool(photo.get("cover", False))),
            int(photo.get("sort_order") or 0),
            photo.get("rights_warning") or "",
            existing["created_at"] if existing else timestamp,
            timestamp,
        ),
    )


def patch_listing(listing_id: str, patch: dict[str, Any], db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    current = get_listing(listing_id, db_path)
    if not current:
        raise KeyError(listing_id)
    current.update(patch)
    if "description" in patch:
        current["description"] = sanitize_public_description(str(patch["description"]), get_settings(db_path).get("forbidden_public_phrases"))
    saved = upsert_listing(current, db_path)
    refresh_validation(listing_id, db_path)
    return saved


def patch_photo(listing_id: str, photo_id: str, patch: dict[str, Any], db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    init_db(db_path)
    with connect(db_path) as conn:
        row = conn.execute("select * from photos where id = ? and listing_id = ?", (photo_id, listing_id)).fetchone()
        if not row:
            raise KeyError(photo_id)
        current = row_to_photo(row)
        current.update(patch)
        if current.get("cover"):
            conn.execute("update photos set cover = 0 where listing_id = ?", (listing_id,))
        upsert_photo(conn, listing_id, current)
    refresh_validation(listing_id, db_path)
    return get_listing(listing_id, db_path) or {}


def delete_listing(listing_id: str, db_path: Path = DEFAULT_DB_PATH) -> bool:
    init_db(db_path)
    with connect(db_path) as conn:
        exists = conn.execute("select 1 from listings where id = ?", (listing_id,)).fetchone()
        if not exists:
            return False
        conn.execute("delete from photos where listing_id = ?", (listing_id,))
        conn.execute("delete from run_log where listing_id = ?", (listing_id,))
        conn.execute("delete from listings where id = ?", (listing_id,))
        return True


def delete_listings(listing_ids: list[str], db_path: Path = DEFAULT_DB_PATH) -> list[str]:
    deleted: list[str] = []
    for listing_id in listing_ids:
        if delete_listing(listing_id, db_path):
            deleted.append(listing_id)
    return deleted


def approve_listing(listing_id: str, approved: bool, db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    listing = get_listing(listing_id, db_path)
    if not listing:
        raise KeyError(listing_id)
    settings = get_settings(db_path)
    issues = validate_listing(listing, settings)
    has_errors = any(issue["severity"] == "error" for issue in issues)
    status = "approved" if approved and not has_errors else "needs_review"
    with connect(db_path) as conn:
        conn.execute(
            "update listings set approved = ?, status = ?, validation_json = ?, updated_at = ? where id = ?",
            (int(approved and not has_errors), status, json.dumps(issues), now(), listing_id),
        )
    return get_listing(listing_id, db_path) or {}


def refresh_validation(listing_id: str, db_path: Path = DEFAULT_DB_PATH) -> None:
    listing = get_listing(listing_id, db_path)
    if not listing:
        return
    issues = validate_listing(listing, get_settings(db_path))
    approved = bool(listing.get("approved")) and not any(issue["severity"] == "error" for issue in issues)
    status = "approved" if approved else listing.get("status") or "needs_review"
    if any(issue["severity"] == "error" for issue in issues) and status == "approved":
        status = "needs_review"
    with connect(db_path) as conn:
        conn.execute(
            "update listings set validation_json = ?, approved = ?, status = ?, updated_at = ? where id = ?",
            (json.dumps(issues), int(approved), status, now(), listing_id),
        )


def revalidate_all(db_path: Path = DEFAULT_DB_PATH) -> None:
    for listing in list_listings(db_path=db_path):
        refresh_validation(listing["id"], db_path)


def export_posting_queue(db_path: Path = DEFAULT_DB_PATH) -> list[dict[str, Any]]:
    settings = get_settings(db_path)
    queue: list[dict[str, Any]] = []
    for listing in list_listings(db_path=db_path):
        if not listing.get("approved"):
            continue
        issues = validate_listing(listing, settings)
        if any(issue["severity"] == "error" for issue in issues):
            continue
        photos = [
            photo
            for photo in sorted(listing["photos"], key=lambda p: (not p["cover"], p["sort_order"], p["id"]))
            if not photo["removed"]
        ]
        queue.append(
            {
                "id": listing["id"],
                "title": listing["title"],
                "price": listing["price"],
                "condition": listing["condition"],
                "category": listing["category"],
                "quantity_text": listing["quantity_text"],
                "description": listing["description"],
                "location": listing["location"] or settings["location"],
                "pickup_enabled": listing["pickup_enabled"],
                "shipping_enabled": listing["shipping_enabled"],
                "package_weight_oz": listing["package_weight_oz"] or settings.get("default_package_weight_oz"),
                "photo_paths": [photo["path"] for photo in photos if photo["path"]],
                "photo_urls": [photo["source_url"] or photo["uri"] for photo in photos if not photo["path"]],
                "auto_publish_allowed": bool(settings.get("auto_publish")),
            }
        )
    return queue


def add_log(level: str, message: str, listing_id: str | None = None, details: dict[str, Any] | None = None, db_path: Path = DEFAULT_DB_PATH) -> None:
    init_db(db_path)
    with connect(db_path) as conn:
        conn.execute(
            "insert into run_log(listing_id, level, message, details_json, created_at) values (?, ?, ?, ?, ?)",
            (listing_id, level, message, json.dumps(details or {}), now()),
        )


def list_logs(limit: int = 100, db_path: Path = DEFAULT_DB_PATH) -> list[dict[str, Any]]:
    init_db(db_path)
    with connect(db_path) as conn:
        rows = conn.execute("select * from run_log order by id desc limit ?", (limit,)).fetchall()
        return [
            {
                "id": row["id"],
                "listing_id": row["listing_id"],
                "level": row["level"],
                "message": row["message"],
                "details": json.loads(row["details_json"] or "{}"),
                "created_at": row["created_at"],
            }
            for row in rows
        ]


BTC_ENTRY_TYPES = {"sale_proceeds", "cash_set_aside", "btc_purchase", "referral_bonus", "adjustment"}


def normalize_btc_entry(data: dict[str, Any]) -> dict[str, Any]:
    entry_type = str(data.get("entry_type") or "sale_proceeds")
    if entry_type not in BTC_ENTRY_TYPES:
        entry_type = "adjustment"
    timestamp = now()
    return {
        "entry_type": entry_type,
        "title": str(data.get("title") or entry_type.replace("_", " ").title())[:150],
        "amount_usd": float(data.get("amount_usd") or 0),
        "btc_amount": float(data.get("btc_amount") or 0),
        "btc_price_usd": float(data["btc_price_usd"]) if data.get("btc_price_usd") not in (None, "") else None,
        "listing_id": str(data.get("listing_id") or "") or None,
        "notes": str(data.get("notes") or ""),
        "entry_date": str(data.get("entry_date") or timestamp[:10]),
    }


def row_to_btc_entry(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "entry_type": row["entry_type"],
        "title": row["title"],
        "amount_usd": row["amount_usd"],
        "btc_amount": row["btc_amount"],
        "btc_price_usd": row["btc_price_usd"],
        "listing_id": row["listing_id"],
        "notes": row["notes"],
        "entry_date": row["entry_date"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def list_btc_entries(db_path: Path = DEFAULT_DB_PATH) -> list[dict[str, Any]]:
    init_db(db_path)
    with connect(db_path) as conn:
        rows = conn.execute("select * from btc_progress_entries order by entry_date desc, id desc").fetchall()
        return [row_to_btc_entry(row) for row in rows]


def create_btc_entry(data: dict[str, Any], db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    init_db(db_path)
    entry = normalize_btc_entry(data)
    timestamp = now()
    with connect(db_path) as conn:
        cursor = conn.execute(
            """
            insert into btc_progress_entries(entry_type, title, amount_usd, btc_amount, btc_price_usd, listing_id, notes, entry_date, created_at, updated_at)
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                entry["entry_type"],
                entry["title"],
                entry["amount_usd"],
                entry["btc_amount"],
                entry["btc_price_usd"],
                entry["listing_id"],
                entry["notes"],
                entry["entry_date"],
                timestamp,
                timestamp,
            ),
        )
        entry_id = cursor.lastrowid
    add_log("info", f"Added BTC progress entry: {entry['title']}", listing_id=entry.get("listing_id"), db_path=db_path)
    return get_btc_entry(int(entry_id), db_path) or {}


def get_btc_entry(entry_id: int, db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any] | None:
    init_db(db_path)
    with connect(db_path) as conn:
        row = conn.execute("select * from btc_progress_entries where id = ?", (entry_id,)).fetchone()
        return row_to_btc_entry(row) if row else None


def patch_btc_entry(entry_id: int, patch: dict[str, Any], db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    current = get_btc_entry(entry_id, db_path)
    if not current:
        raise KeyError(entry_id)
    current.update(patch)
    entry = normalize_btc_entry(current)
    with connect(db_path) as conn:
        conn.execute(
            """
            update btc_progress_entries
            set entry_type = ?, title = ?, amount_usd = ?, btc_amount = ?, btc_price_usd = ?, listing_id = ?, notes = ?, entry_date = ?, updated_at = ?
            where id = ?
            """,
            (
                entry["entry_type"],
                entry["title"],
                entry["amount_usd"],
                entry["btc_amount"],
                entry["btc_price_usd"],
                entry["listing_id"],
                entry["notes"],
                entry["entry_date"],
                now(),
                entry_id,
            ),
        )
    return get_btc_entry(entry_id, db_path) or {}


def delete_btc_entry(entry_id: int, db_path: Path = DEFAULT_DB_PATH) -> bool:
    init_db(db_path)
    with connect(db_path) as conn:
        cursor = conn.execute("delete from btc_progress_entries where id = ?", (entry_id,))
        return cursor.rowcount > 0


def btc_progress_summary(db_path: Path = DEFAULT_DB_PATH) -> dict[str, Any]:
    settings = get_settings(db_path)
    entries = list_btc_entries(db_path)
    btc_goal = float(settings.get("btc_goal_amount") or 1)
    starting_btc = float(settings.get("btc_owned") or 0)
    manual_price = float(settings.get("manual_btc_usd_price") or 0)
    proceeds_usd = 0.0
    spent_on_btc_usd = 0.0
    ledger_btc = 0.0
    for entry in entries:
        entry_type = entry["entry_type"]
        amount_usd = float(entry.get("amount_usd") or 0)
        btc_amount = float(entry.get("btc_amount") or 0)
        if entry_type == "btc_purchase":
            spent_on_btc_usd += amount_usd
            ledger_btc += btc_amount
        elif entry_type in {"sale_proceeds", "cash_set_aside", "referral_bonus", "adjustment"}:
            proceeds_usd += amount_usd
            ledger_btc += btc_amount
    available_usd = proceeds_usd - spent_on_btc_usd
    estimated_purchasable_btc = max(available_usd, 0) / manual_price if manual_price > 0 else 0
    total_btc_owned = starting_btc + ledger_btc
    projected_btc = total_btc_owned + estimated_purchasable_btc
    remaining_btc = max(btc_goal - projected_btc, 0)
    return {
        "goal_btc": btc_goal,
        "starting_btc_owned": starting_btc,
        "ledger_btc": ledger_btc,
        "total_btc_owned": total_btc_owned,
        "manual_btc_usd_price": manual_price,
        "gross_proceeds_usd": proceeds_usd,
        "spent_on_btc_usd": spent_on_btc_usd,
        "available_usd": available_usd,
        "estimated_purchasable_btc": estimated_purchasable_btc,
        "projected_btc": projected_btc,
        "remaining_btc": remaining_btc,
        "remaining_usd": remaining_btc * manual_price if manual_price > 0 else None,
        "progress_percent": min((projected_btc / btc_goal) * 100, 100) if btc_goal > 0 else 0,
        "entry_count": len(entries),
        "currency": settings.get("progress_currency") or "USD",
        "google_sheet_url": settings.get("google_sheet_url") or "",
        "kraken_referral_url": settings.get("kraken_referral_url") or "",
        "recent_entries": entries[:5],
    }


def btc_entries_csv(db_path: Path = DEFAULT_DB_PATH) -> str:
    import csv

    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["id", "entry_date", "entry_type", "title", "amount_usd", "btc_amount", "btc_price_usd", "listing_id", "notes"],
    )
    writer.writeheader()
    for entry in reversed(list_btc_entries(db_path)):
        writer.writerow({key: entry.get(key) for key in writer.fieldnames})
    return output.getvalue()


def btc_entries_xlsx(db_path: Path = DEFAULT_DB_PATH) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BTC Progress Ledger"
    headers = ["ID", "Date", "Type", "Title", "Amount USD", "BTC Amount", "BTC Price USD", "Listing ID", "Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
    for entry in reversed(list_btc_entries(db_path)):
        ws.append(
            [
                entry["id"],
                entry["entry_date"],
                entry["entry_type"],
                entry["title"],
                entry["amount_usd"],
                entry["btc_amount"],
                entry["btc_price_usd"],
                entry["listing_id"],
                entry["notes"],
            ]
        )
    widths = [10, 14, 18, 32, 14, 14, 16, 18, 48]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = "A2"
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
